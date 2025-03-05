from openpyxl import Workbook
from openpyxl.styles import Protection

# Buat workbook dan worksheet
wb = Workbook()
ws = wb.active

# Contoh data (5 baris di kolom A)
data = ["Data1", "Data2", "Data3", "Data4", "Data5"]
for i, value in enumerate(data, start=1):
    ws[f"A{i}"] = value  # Isi kolom A
    ws[f"C{i}"] = f"Editable {i}"  # Isi kolom C

# **1. Lindungi seluruh sheet**
ws.protection.sheet = True  # Aktifkan proteksi sheet

# **2. Cari jumlah baris yang terisi di kolom A**
max_row = len(data)  # Bisa juga pakai: ws.max_row kalau ada banyak data

# **3. Unprotect kolom C sesuai jumlah baris di kolom A**
for row in range(1, max_row + 1):  # Loop dari C1 sampai Cn
    ws[f"C{row}"].protection = Protection(locked=False)  # Buka proteksi

# Simpan file
wb.save("protected_dynamic.xlsx")
